import re
import uuid
import openpyxl
from django.db import transaction
from django.utils import timezone
from django.db.models import Q
from core.audit import log_action
from beneficiaries.models import Region, Site, Employee, Dependent

class EmployeeService:
    @staticmethod
    @transaction.atomic
    def create_employee(user, employee_number, last_name, first_name, site_id, post_name=None, address=None, employment_status='ACTIVE'):
        """
        Creates a new employee and logs the transaction.
        """
        site = Site.objects.get(pk=site_id)
        
        # Check duplicate employee number
        if Employee.objects.filter(employee_number=employee_number).exists():
            raise ValueError(f"Employee number '{employee_number}' is already in use.")

        employee = Employee.objects.create(
            employee_number=employee_number,
            last_name=last_name.upper().strip(),
            post_name=post_name.strip() if post_name else None,
            first_name=first_name.strip(),
            site=site,
            address=address,
            employment_status=employment_status
        )
        
        log_action(
            user=user,
            action='CREATE',
            instance=employee,
            changes={
                "employee_number": [None, employee.employee_number],
                "name": [None, f"{employee.last_name} {employee.first_name}"],
                "site": [None, site.name]
            }
        )
        return employee

    @staticmethod
    @transaction.atomic
    def update_employee(user, employee_id, **fields):
        """
        Updates an employee instance and logs the audit trail.
        """
        employee = Employee.objects.get(pk=employee_id)
        changes = {}
        
        for field, value in fields.items():
            if hasattr(employee, field):
                old_val = getattr(employee, field)
                
                # Special styling constraints
                if field == 'last_name' and value:
                    value = value.upper().strip()
                elif field in ['first_name', 'post_name'] and value:
                    value = value.strip()
                
                if old_val != value:
                    setattr(employee, field, value)
                    changes[field] = [str(old_val), str(value)]
                    
        if changes:
            employee.save()
            log_action(user=user, action='UPDATE', instance=employee, changes=changes)
            
        return employee

    @staticmethod
    @transaction.atomic
    def delete_employee(user, employee_id):
        """
        Soft-deletes an employee and logs the transaction.
        """
        employee = Employee.objects.get(pk=employee_id)
        employee.delete()  # Performs soft delete
        
        log_action(
            user=user,
            action='DELETE',
            instance=employee,
            changes={"is_deleted": [False, True]}
        )
        return employee


class DependentService:
    @staticmethod
    @transaction.atomic
    def add_dependent(user, employee_id, full_name, gender, relationship, birth_date=None):
        """
        Adds a dependent to an employee.
        """
        employee = Employee.objects.get(pk=employee_id)
        
        # Check duplicate dependent per employee
        if Dependent.objects.filter(employee=employee, full_name=full_name.strip(), relationship=relationship).exists():
            raise ValueError(f"Dependent '{full_name}' is already registered as a {relationship.lower()} for this employee.")
            
        dependent = Dependent.objects.create(
            employee=employee,
            full_name=full_name.strip(),
            gender=gender,
            relationship=relationship,
            birth_date=birth_date
        )
        
        log_action(
            user=user,
            action='CREATE',
            instance=dependent,
            changes={
                "employee": [None, employee.employee_number],
                "full_name": [None, dependent.full_name],
                "relationship": [None, dependent.relationship]
            }
        )
        return dependent


class ExcelImportService:
    @staticmethod
    def _parse_names(full_name_str):
        """
        Parses combined 'Post-nom et Prénom' into post_name and first_name.
        Assumes first token is post_name, subsequent tokens are first_name.
        """
        if not full_name_str:
            return "", ""
        
        tokens = [t.strip() for t in str(full_name_str).split() if t.strip()]
        if not tokens:
            return "", ""
        
        post_name = tokens[0]
        first_name = " ".join(tokens[1:]) if len(tokens) > 1 else ""
        return post_name, first_name

    @staticmethod
    def _generate_employee_number(sequence):
        """
        Generates a standardized employee number if missing.
        Format: EMP-2026-{sequence:04d}
        """
        current_year = timezone.now().year
        return f"EMP-{current_year}-{sequence:04d}"

    @classmethod
    @transaction.atomic
    def import_excel(cls, user, file_file):
        """
        Imports beneficiaries from an Excel file.
        Each worksheet represents a Region.
        """
        workbook = openpyxl.load_workbook(file_file, data_only=True)
        report = {
            "success": True,
            "summary": {
                "regions_created": 0,
                "sites_created": 0,
                "employees_created": 0,
                "employees_updated": 0,
                "dependents_created": 0,
                "errors_count": 0,
                "warnings_count": 0
            },
            "sheets_processed": [],
            "logs": []
        }
        
        emp_sequence = Employee.objects.count() + 1
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            region_name = sheet_name.strip()
            
            if not region_name or sheet.max_row <= 1:
                continue
                
            # Auto-create Region
            region, region_created = Region.objects.get_or_create(name=region_name)
            if region_created:
                report["summary"]["regions_created"] += 1
                report["logs"].append({
                    "level": "INFO",
                    "sheet": region_name,
                    "row": None,
                    "message": f"Region '{region_name}' created automatically."
                })
                
            # Auto-create default site inside this region
            site_name = "Centrale"
            site, site_created = Site.objects.get_or_create(region=region, name=site_name)
            if site_created:
                report["summary"]["sites_created"] += 1
                
            sheet_info = {
                "name": region_name,
                "rows_processed": 0,
                "employees_created": 0,
                "dependents_created": 0
            }
            
            current_employee = None
            
            # Read header
            headers = [cell.value for cell in sheet[1]]
            
            # Map column names to index numbers
            # Target columns: N, Nom, Post-nom et Prénom, Épouse, Nom de l'enfant, Année de naissance, Sexe, Addresse
            col_map = {}
            for idx, h in enumerate(headers):
                if not h:
                    continue
                h_clean = str(h).strip().lower()
                if 'n' == h_clean or 'n°' in h_clean:
                    col_map['n'] = idx
                elif 'nom' == h_clean:
                    col_map['nom'] = idx
                elif 'post-nom' in h_clean or 'prenom' in h_clean or 'prénom' in h_clean:
                    col_map['post_prenom'] = idx
                elif 'epouse' in h_clean or 'épouse' in h_clean:
                    col_map['epouse'] = idx
                elif 'enfant' in h_clean:
                    col_map['enfant'] = idx
                elif 'naissance' in h_clean or 'annee' in h_clean or 'année' in h_clean:
                    col_map['naissance'] = idx
                elif 'sexe' in h_clean or 'gender' in h_clean:
                    col_map['sexe'] = idx
                elif 'addresse' in h_clean or 'adresse' in h_clean:
                    col_map['adresse'] = idx

            # Fallback mappings if columns didn't match perfectly
            if 'nom' not in col_map and len(headers) > 1:
                col_map['nom'] = 1
            if 'post_prenom' not in col_map and len(headers) > 2:
                col_map['post_prenom'] = 2
            if 'epouse' not in col_map and len(headers) > 3:
                col_map['epouse'] = 3
            if 'enfant' not in col_map and len(headers) > 4:
                col_map['enfant'] = 4
            if 'naissance' not in col_map and len(headers) > 5:
                col_map['naissance'] = 5
            if 'sexe' not in col_map and len(headers) > 6:
                col_map['sexe'] = 6
            if 'adresse' not in col_map and len(headers) > 7:
                col_map['adresse'] = 7

            for r_idx in range(2, sheet.max_row + 1):
                row = [cell.value for cell in sheet[r_idx]]
                
                # Skip completely blank rows
                if not any(row):
                    continue
                    
                sheet_info["rows_processed"] += 1
                
                # Fetch row cell values safely
                n_val = str(row[col_map['n']]).strip() if col_map.get('n') is not None and row[col_map['n']] is not None else ""
                nom_val = str(row[col_map['nom']]).strip() if col_map.get('nom') is not None and row[col_map['nom']] is not None else ""
                post_prenom_val = str(row[col_map['post_prenom']]).strip() if col_map.get('post_prenom') is not None and row[col_map['post_prenom']] is not None else ""
                epouse_val = str(row[col_map['epouse']]).strip() if col_map.get('epouse') is not None and row[col_map['epouse']] is not None else ""
                enfant_val = str(row[col_map['enfant']]).strip() if col_map.get('enfant') is not None and row[col_map['enfant']] is not None else ""
                naissance_val = str(row[col_map['naissance']]).strip() if col_map.get('naissance') is not None and row[col_map['naissance']] is not None else ""
                sexe_val = str(row[col_map['sexe']]).strip().upper() if col_map.get('sexe') is not None and row[col_map['sexe']] is not None else ""
                adresse_val = str(row[col_map['adresse']]).strip() if col_map.get('adresse') is not None and row[col_map['adresse']] is not None else ""

                # --- 1. PROCESS EMPLOYEE ---
                if nom_val or post_prenom_val:
                    post_name, first_name = cls._parse_names(post_prenom_val)
                    
                    # Deduplicate Employee by name within region/site
                    existing_employees = Employee.objects.filter(
                        last_name=nom_val.upper(),
                        post_name=post_name if post_name else None,
                        first_name=first_name,
                        site__region=region
                    )
                    
                    if existing_employees.exists():
                        current_employee = existing_employees.first()
                        report["summary"]["employees_updated"] += 1
                        report["logs"].append({
                            "level": "WARNING",
                            "sheet": region_name,
                            "row": r_idx,
                            "message": f"Employee '{nom_val} {post_prenom_val}' already exists. Appending dependents."
                        })
                        report["summary"]["warnings_count"] += 1
                    else:
                        # Auto generate employee number if missing or numeric index
                        employee_number = n_val if n_val and not n_val.isdigit() else cls._generate_employee_number(emp_sequence)
                        emp_sequence += 1
                        
                        try:
                            current_employee = Employee.objects.create(
                                employee_number=employee_number,
                                last_name=nom_val.upper(),
                                post_name=post_name if post_name else None,
                                first_name=first_name,
                                site=site,
                                address=adresse_val
                            )
                            sheet_info["employees_created"] += 1
                            report["summary"]["employees_created"] += 1
                            
                            # Log audit actions
                            log_action(
                                user=user,
                                action='CREATE',
                                instance=current_employee,
                                changes={
                                    "employee_number": [None, current_employee.employee_number],
                                    "name": [None, f"{current_employee.last_name} {current_employee.first_name}"],
                                    "site": [None, site.name]
                                }
                            )
                        except Exception as e:
                            report["summary"]["errors_count"] += 1
                            report["logs"].append({
                                "level": "ERROR",
                                "sheet": region_name,
                                "row": r_idx,
                                "message": f"Failed to create employee '{nom_val}': {str(e)}"
                            })
                            current_employee = None
                            continue

                # --- 2. PROCESS SPOUSE ---
                if current_employee and epouse_val and epouse_val.lower() not in ['-', 'none', 'n/a', '']:
                    # Separate multiple spouses if any (e.g. split by comma)
                    spouses = [s.strip() for s in re.split(r'[,&/]', epouse_val) if s.strip()]
                    for spouse_name in spouses:
                        if spouse_name.lower() in ['-', 'none', 'n/a', '']:
                            continue
                        
                        # Add Spouse dependent
                        if not Dependent.objects.filter(employee=current_employee, full_name=spouse_name, relationship='SPOUSE').exists():
                            try:
                                dep = Dependent.objects.create(
                                    employee=current_employee,
                                    full_name=spouse_name,
                                    gender='F',  # Default gender for wives, can be corrected manually
                                    relationship='SPOUSE'
                                )
                                sheet_info["dependents_created"] += 1
                                report["summary"]["dependents_created"] += 1
                                
                                log_action(
                                    user=user,
                                    action='CREATE',
                                    instance=dep,
                                    changes={"employee": [None, current_employee.employee_number], "full_name": [None, spouse_name], "relationship": [None, "SPOUSE"]}
                                )
                            except Exception as e:
                                report["summary"]["errors_count"] += 1
                                report["logs"].append({
                                    "level": "ERROR",
                                    "sheet": region_name,
                                    "row": r_idx,
                                    "message": f"Failed to add spouse '{spouse_name}' to employee '{current_employee}': {str(e)}"
                                })

                # --- 3. PROCESS CHILD ---
                if current_employee and enfant_val and enfant_val.lower() not in ['-', 'none', 'n/a', '']:
                    # Clean and format gender
                    gender_mapped = 'M'
                    if sexe_val in ['F', 'FEMININ', 'FÉMININ', 'FILLE', 'FEMELLE']:
                        gender_mapped = 'F'
                    elif sexe_val in ['M', 'MASCULIN', 'GARÇON', 'GARCON', 'MALE']:
                        gender_mapped = 'M'
                    else:
                        if sexe_val:
                            report["logs"].append({
                                "level": "WARNING",
                                "sheet": region_name,
                                "row": r_idx,
                                "message": f"Unrecognized gender '{sexe_val}' for child '{enfant_val}'. Defaulted to 'M'."
                            })
                            report["summary"]["warnings_count"] += 1

                    # Parse Birth Year
                    birth_date = None
                    if naissance_val:
                        # Extract 4-digit number
                        year_match = re.search(r'\b(19\d\d|20\d\d)\b', str(naissance_val))
                        if year_match:
                            birth_date = f"{year_match.group(1)}-01-01"
                        else:
                            report["logs"].append({
                                "level": "WARNING",
                                "sheet": region_name,
                                "row": r_idx,
                                "message": f"Could not parse birth year '{naissance_val}' for child '{enfant_val}'."
                            })
                            report["summary"]["warnings_count"] += 1

                    # Create child
                    if not Dependent.objects.filter(employee=current_employee, full_name=enfant_val, relationship='CHILD').exists():
                        try:
                            dep = Dependent.objects.create(
                                employee=current_employee,
                                full_name=enfant_val,
                                gender=gender_mapped,
                                relationship='CHILD',
                                birth_date=birth_date
                            )
                            sheet_info["dependents_created"] += 1
                            report["summary"]["dependents_created"] += 1
                            
                            log_action(
                                user=user,
                                action='CREATE',
                                instance=dep,
                                changes={"employee": [None, current_employee.employee_number], "full_name": [None, enfant_val], "relationship": [None, "CHILD"]}
                            )
                        except Exception as e:
                            report["summary"]["errors_count"] += 1
                            report["logs"].append({
                                "level": "ERROR",
                                "sheet": region_name,
                                "row": r_idx,
                                "message": f"Failed to add child '{enfant_val}' to employee '{current_employee}': {str(e)}"
                            })

            report["sheets_processed"].append(sheet_info)

        # Log import job action
        if report["summary"]["errors_count"] > 0 and report["summary"]["employees_created"] == 0:
            report["success"] = False

        return report
