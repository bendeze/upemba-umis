import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill
from django.core.management.base import BaseCommand
from pharmacy.models import PharmacyStock, StockMovement, MedicineBatch

class Command(BaseCommand):
    help = 'Exports all pharmacy data to an Excel file on the Desktop'

    def handle(self, *args, **options):
        desktop_path = os.path.join(str(Path.home()), 'Desktop', 'UMIS_Pharmacy_Backups')
        os.makedirs(desktop_path, exist_ok=True)
        
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Pharmacy_Backup_{timestamp}.xlsx"
        filepath = os.path.join(desktop_path, filename)

        wb = openpyxl.Workbook()
        
        # Sheet 1: Current Stock
        ws_stock = wb.active
        ws_stock.title = "Current Stock"
        headers_stock = ["Site", "Medicine Name", "Reference Number", "Quantity", "Unit"]
        ws_stock.append(headers_stock)
        
        for stock in PharmacyStock.objects.select_related('medical_center', 'medicine').all():
            ws_stock.append([
                stock.medical_center.name if stock.medical_center else '',
                stock.medicine.name,
                stock.medicine.reference_number or '',
                stock.quantity,
                stock.medicine.unit
            ])

        # Sheet 2: Stock Movements
        ws_movements = wb.create_sheet(title="Stock Movements")
        headers_move = ["Date", "Site", "Movement Type", "Medicine Name", "Quantity", "Expiration Date", "Notes"]
        ws_movements.append(headers_move)
        
        for move in StockMovement.objects.select_related('medical_center', 'medicine').all():
            ws_movements.append([
                move.date.strftime("%Y-%m-%d"),
                move.medical_center.name if move.medical_center else '',
                move.movement_type,
                move.medicine.name,
                move.quantity,
                move.expiration_date.strftime("%Y-%m-%d") if move.expiration_date else '',
                move.notes or ''
            ])

        # Sheet 3: Medicine Batches (Expirations)
        ws_batches = wb.create_sheet(title="Active Batches")
        headers_batches = ["Site", "Medicine Name", "Expiration Date", "Remaining Quantity"]
        ws_batches.append(headers_batches)
        
        for batch in MedicineBatch.objects.select_related('medical_center', 'medicine').filter(quantity__gt=0):
            ws_batches.append([
                batch.medical_center.name if batch.medical_center else '',
                batch.medicine.name,
                batch.expiration_date.strftime("%Y-%m-%d") if batch.expiration_date else '',
                batch.quantity
            ])

        # Formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
        
        for sheet in wb.worksheets:
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill

        wb.save(filepath)
        self.stdout.write(self.style.SUCCESS(f"Successfully exported pharmacy backup to {filepath}"))
