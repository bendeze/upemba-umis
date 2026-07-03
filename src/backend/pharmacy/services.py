import openpyxl
from django.db import transaction
from pharmacy.models import Medicine, StockMovement

class PharmacyExcelService:
    @staticmethod
    @transaction.atomic
    def import_requisitions(medical_center, file_obj):
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb.active
        
        # Find column indices
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h).strip().lower() if h else '' for h in header_row]
        
        try:
            name_idx = next(i for i, h in enumerate(headers) if 'name' in h or 'nom' in h)
            qty_idx = next(i for i, h in enumerate(headers) if 'quantit' in h or 'qty' in h)
        except StopIteration:
            raise ValueError("Missing required columns: Medicine Name or Quantity.")
            
        ref_idx = next((i for i, h in enumerate(headers) if 'ref' in h or 'lot' in h), None)
        unit_idx = next((i for i, h in enumerate(headers) if 'unit' in h), None)
        exp_idx = next((i for i, h in enumerate(headers) if 'exp' in h or 'date' in h), None)

        movements_created = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[name_idx] or not row[qty_idx]:
                continue
                
            med_name = str(row[name_idx]).strip()
            try:
                quantity = int(row[qty_idx])
            except (ValueError, TypeError):
                continue # Skip invalid quantity
                
            if quantity <= 0:
                continue
                
            ref = str(row[ref_idx]).strip() if ref_idx is not None and row[ref_idx] else None
            unit = str(row[unit_idx]).strip() if unit_idx is not None and row[unit_idx] else 'Unité'
            exp_date = row[exp_idx] if exp_idx is not None else None
            if hasattr(exp_date, 'date'):
                exp_date = exp_date.date()

            # Get or Create Medicine
            medicine, created = Medicine.objects.get_or_create(
                name=med_name,
                defaults={
                    'unit': unit
                }
            )

            # Create Stock Movement (IN)
            StockMovement.objects.create(
                medical_center=medical_center,
                medicine=medicine,
                movement_type='IN',
                quantity=quantity,
                lot_number=ref if ref else 'UNKNOWN',
                expiration_date=exp_date,
                notes='Imported from Excel Requisition'
            )
            movements_created += 1
            
        return movements_created

    @staticmethod
    @transaction.atomic
    def import_consumptions(medical_center, file_obj):
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb.active
        
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h).strip().lower() if h else '' for h in header_row]
        
        try:
            name_idx = next(i for i, h in enumerate(headers) if 'name' in h or 'nom' in h)
            qty_idx = next(i for i, h in enumerate(headers) if 'quantit' in h or 'qty' in h)
        except StopIteration:
            raise ValueError("Missing required columns: Medicine Name or Quantity.")

        movements_created = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[name_idx] or not row[qty_idx]:
                continue
                
            med_name = str(row[name_idx]).strip()
            try:
                quantity = int(row[qty_idx])
            except (ValueError, TypeError):
                continue
                
            if quantity <= 0:
                continue

            # For OUT movement, we expect the medicine to exist
            try:
                medicine = Medicine.objects.get(name__iexact=med_name)
            except Medicine.DoesNotExist:
                continue # Skip if medicine not found

            # Create Stock Movement (OUT)
            StockMovement.objects.create(
                medical_center=medical_center,
                medicine=medicine,
                movement_type='OUT',
                quantity=quantity,
                notes='Imported from Excel Consumption'
            )
            movements_created += 1

        return movements_created

    @staticmethod
    @transaction.atomic
    def dispense_prescription(prescription, item_quantities, user_notes=''):
        '''
        item_quantities is a dict mapping prescription_item_id to quantity_to_dispense
        '''
        from pharmacy.models import StockMovement, PrescriptionItem
        
        movements = []
        for item_id, qty in item_quantities.items():
            if qty <= 0:
                continue
            
            try:
                item = PrescriptionItem.objects.get(id=item_id, prescription=prescription)
            except PrescriptionItem.DoesNotExist:
                continue
                
            # Create StockMovement
            movement = StockMovement.objects.create(
                medical_center=prescription.medical_center,
                medicine=item.medicine,
                movement_type='DISPENSE',
                quantity=qty,
                prescription_item=item,
                notes=user_notes
            )
            movements.append(movement)
            
            # Update item dispensed quantity
            item.quantity_dispensed += qty
            item.save()
            
        # Update prescription status
        all_items = prescription.items.all()
        total_prescribed = sum(i.quantity_prescribed for i in all_items)
        total_dispensed = sum(i.quantity_dispensed for i in all_items)
        
        if total_dispensed >= total_prescribed:
            prescription.status = 'COMPLETED'
        elif total_dispensed > 0:
            prescription.status = 'PARTIAL'
            
        prescription.save()
        
        return movements
